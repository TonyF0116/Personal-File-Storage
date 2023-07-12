from flask import Blueprint, url_for, request, send_file
from ..utils.jwt_validation import jwt_validation
from ..models.edit import get_file_info, update_file_modify_time
from ..models.index import add_new_file, check_belonging
import pandas as pd
from datetime import datetime
from ..config import sender_email, smtp_server, smtp_port, email_username, email_password

from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Handle requests from the edit page
blueprint = Blueprint('edit', __name__, url_prefix='/api/edit')


@blueprint.route('', methods=['POST'])
def edit():
    # Retrieve the token from request header
    # Return redirection if Authorization not in header
    authorization = request.headers.get('Authorization')
    if authorization == None:
        return {'msg': 'Validation failed',
                'data': {'redirection': '{}?redirection={}&warning=Login+Expired'
                         .format(url_for('route_account'), url_for('route_edit'))}}, 401

    # Decode the JWT token
    token = authorization[7:]
    result = jwt_validation(token, url_for('route_edit'))

    # Validation failed => Invalid token => Login again
    if result['msg'] == 'Validation failed':
        return {'msg': 'Validation failed',
                'data': {'redirection': '{}?redirection={}&warning=Login+Expired'
                         .format(url_for('route_account'), url_for('route_edit'))}}, 401

    # Unauthorized => Valid token with no auth => Go to account page to get authorization
    if result['msg'] == 'Unauthorized':
        return {'msg': "Unauthorized",
                'data': {'redirection': '{}?redirection={}&file_id={}&Authorization={}'
                         .format(url_for('route_account'), url_for('route_edit'), request.args.get('file_id'), authorization)}}, 302

    account_id = result['data']['payload']['account_id']

    # Retrieve file type and build return info
    file_id = request.args.get('file_id')
    file_type = get_file_info(file_id)[0][2]

    return {'msg': None,
            'data': {'account_id': account_id,
                     'file_type': file_type}}, 200


# Routes for serving files for the edit page
@blueprint.route('/get_file')
def get_file():
    # Retrieve info from the args
    account_id = request.args.get('account_id')
    file_id = request.args.get('file_id')

    # If file id not given, return warning png
    if file_id == 'undefined':
        return send_file('files/Unauthorized.png')

    file_info = get_file_info(file_id)

    # Check belonging
    if int(file_info[0][0]) != int(account_id):
        if file_info[0][2] == 0:
            return send_file('files/Unauthorized.png')
        elif file_info[0][2] == 1:
            return send_file('files/Unauthorized.pdf')
        else:
            pass
    else:
        return send_file('files/{}/{}'.format(file_info[0][0], file_info[0][1]))


# Route for updating excel files
@blueprint.route('/save_excel', methods=['POST'])
def save_excel():
    # Retrieve the file id and updated data from the request
    excel_data = request.get_json()
    file_id = request.args.get('file_id')

    # Get belonging and file name from the database
    file_info = get_file_info(file_id)
    account_id = file_info[0][0]
    file_name = file_info[0][1]

    # Construct file path
    file_path = 'server/files/{}/{}'.format(account_id, file_name)

    # Do the update
    try:
        df = pd.read_excel(file_path, header=None)
        max_rows = max(len(df), len(excel_data))
        max_cols = max(len(df.columns), max(len(row) for row in excel_data))
        df = df.reindex(index=range(max_rows), columns=range(max_cols))

        for row_index, row in enumerate(excel_data):
            for cell_index, value in enumerate(row):
                df.iat[row_index, cell_index] = value

        df.to_excel(file_path, header=False, index=False)

        # Update last modified time
        update_file_modify_time(
            file_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        return {'msg': 'Save successful',
                'data': None}, 200

    except Exception as error:
        print(error)
        return {'msg': 'Save Failed. Check server terminal for more info.',
                'data': None}, 500


# Route for generating pdf from excel
@blueprint.route('/excel_to_pdf', methods=['POST'])
def excel_to_pdf():
    # Retrieve file id from request args
    file_id = request.args.get('file_id')

    # Get belonging and file name from the database
    file_info = get_file_info(file_id)
    account_id = file_info[0][0]
    file_name = file_info[0][1]

    try:

        # Load excel file
        sheet = load_workbook(
            'server/files/{}/{}'.format(account_id, file_name)).active

        # Construct new file name
        new_file_name_tmp = file_name.split('.')
        new_file_name = ''.join(new_file_name_tmp[0:len(new_file_name_tmp)-1])

        # Create pdf file
        pdf = canvas.Canvas(
            'server/files/{}/{}.pdf'.format(account_id, new_file_name), pagesize=letter)

        pdf.setFont("Helvetica", 12)
        y = 750

        # Iterate through each row and column in the sheet
        for row in sheet.iter_rows(values_only=True):
            x = 50
            for cell in row:
                # Add the cell value to the PDF
                pdf.drawString(x, y, str(cell))
                x += 100
            y -= 20

        # Save the PDF file
        pdf.save()

        add_new_file(account_id, '{}.pdf'.format(new_file_name),
                     1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        return {'msg': 'Build successful',
                'data': None}, 200

    except Exception as error:
        print(error)
        return {'msg': 'Build Failed. Check server terminal for more info.',
                'data': None}, 500


# Route for email file sharing
@blueprint.route('/share', methods=['POST'])
def share():
    # Retrieve the email, authorization token and file_id from request args
    authorization = request.args.get('Authorization')
    file_id = request.args.get('file_id')
    email = request.args.get('email')

    # Decode the JWT token
    token = authorization[7:]
    result = jwt_validation(token, url_for('route_index'))

    # Validation failed => Invalid token => Login again
    if result['msg'] == 'Validation failed':
        return {'msg': 'Validation failed',
                'data': {'redirection': '{}?redirection={}&warning=Login+Expired'
                         .format(url_for('route_account'), url_for('route_index'))}}, 401

    # Retrieve account_id and check belonging
    account_id = result['data']['payload']['account_id']
    result = check_belonging(account_id, file_id)
    if len(result) == 1:
        file_path = 'server/files/{}/{}'.format(account_id, result[0][0])
        file_name = result[0][0]

        subject = 'PFS system email'
        message = 'User {} shared a file with you.'.format(account_id)

        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = email
        msg['Subject'] = subject
        msg.attach(MIMEText(message, 'plain'))

        attachment = MIMEBase('application', 'octet-stream')
        with open(file_path, 'rb') as f:
            attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition',
                              f'attachment; filename="{file_name}"')
        msg.attach(attachment)

        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.login(email_username, email_password)
                server.sendmail(sender_email, email, msg.as_string())
        except Exception as error:
            print(error)
            return {'msg': 'Share Failed. Check server terminal for more info.',
                    'data': None}, 500

        return {'msg': "Success",
                'data': None}, 200
    else:
        return {'msg': "Unauthorized",
                'data': None}, 401
